from src.database.database import db
from src.excel.query_sql.sql_recapAll import sql_recapAll
from openpyxl import Workbook
import smtplib
from smtplib import *
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

def main():



    # Create Excel (.xlsx) file -----------------------------------------------
    wb = Workbook()
        # Connect to DB -----------------------------------------------------------
    cur = db.cursor()

    cur.execute(sql_recapAll)
    results = cur.fetchall()
    ws = wb.create_sheet(0)
    ws.title = 'MOVE IN'
    for row in results:
        ws.append(row)
    ws["A17"] = "=SUM(A5:A6)"

    workbook_name = "test_workbook"
    wb.save(workbook_name + ".xlsx")

    fromaddr = "report@autodkms.com"
    toaddr=['oktoberlin@gmail.com']
    cc=['linibelajar@gmail.com']
    
    msg = MIMEMultipart()

    msg['From'] = 'Report DKM Lampung <report@autodkms.com>'
    msg['To'] = ", ".join(toaddr)
    msg['CC'] = ", ".join(cc)
    msg['Subject'] = f"DKM Lampung- REPORT OTOMATIS"

    body = f"Auto Generate Excel Report "

    msg.attach(MIMEText(body, 'plain'))

    filename = f'{workbook_name}.xlsx'
    attachment = open(filename, "rb")

    part = MIMEBase('application','octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

    msg.attach(part)

    server = smtplib.SMTP('smtp.gmail.com', 587)
    #server.set_debuglevel(3)
    server.starttls()
    server.login("linibelajar@gmail.com", "pflugwrmuymomkis")
    text = msg.as_string()
    server.sendmail(fromaddr, toaddr+cc, text)

    print('email successfully sent')
    server.quit()
    

if  __name__ =='__main__':main() 